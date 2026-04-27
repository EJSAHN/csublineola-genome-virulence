from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def write_excel_workbook(sheet_mapping: list[tuple[str, pd.DataFrame]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, dataframe in sheet_mapping:
            dataframe.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    workbook = load_workbook(output_path)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for column_cells in worksheet.columns:
            values = [str(cell.value) for cell in column_cells if cell.value is not None]
            width = max((len(value) for value in values), default=0) + 2
            width = min(width, 40)
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = width

    workbook.save(output_path)
