from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            max_length = 0
            for cell in column_cells[:5000]:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 60))
            width = min(max(max_length + 2, 10), 42)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
    workbook.save(path)


def write_output_workbook(sheets: list[tuple[str, pd.DataFrame]], output: str | Path) -> None:
    output_path = Path(output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, data in sheets:
            data.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    _format_workbook(output_path)
